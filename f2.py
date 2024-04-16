import openpyxl
import pywhatkit
import datetime

# Load the Excel workbook
workbook = openpyxl.load_workbook('Attendance.xlsx')

# Access the first worksheet
sheet = workbook.active

# Function to get phone number based on roll number
def get_phone_number(roll_number):
    # Your phone number mapping
    phone_number_mapping = {
        # Your phone number mappings here
        1: "+919967731898",
        2: "+918888900474",
        3: "+919975970806",
        4: "+919356387615",
        5: "+918767134620",
        6: "+919503583713",
        7: "+918421129248",
        8: "+918767012814",
        9: "+918080369251",
        10:"+918261825587",
        11:"+917774849525",
        12:"+918446379837",
        13:"+918698423639",
        14:"+917758026483",
        15:"+919422506788",
        16:"+917249568883",
        17:"+919404859780",
        18:"+919604288503",
        19:"+919529142864",
        20:"+919322815269",
        21:"+918767356275",
        22:"+919370590582",
        23:"+919841869497",
        24:"+918767057191",
        25:"+918828617839",
        26:"+919284154923",
        27:"+917559128254",
        28:"+919960741427",
        29:"+918805682816",
        30:"+918983801956",
        31:"+919082313113",
        32:"+918999925311",
        33:"+919021784474",
        34:"+918010220830",
        35:"+919834830159",
        36:"+917972158376",
        37:"+919309954203",
        38:"+918626033684",
        39:"+919096278236",
        40:"+919356990714",
        41:"+917507278821",
        42:"+918010200402",
        43:"+919518717159",
        44:"+916354647752",
        45:"+919404205258",
        46:"+918010812171",
        47:"+919503357131",
        48:"+917378964312",
        49:"+919960853540",
        50:"+918624984442",
        51:"+918862022554",
        52:"+917972006141",
        53:"+919730213645",
        54:"+917843093261",
        55:"+919321445801",
        56:"+919764896671",
        57:"+917588529503",
        58:"+918767276942",
        59:"+919322433516",
        60:"+917499558321",
        61:"+919021858988",
        62:"+918080208864",
    }
    return phone_number_mapping.get(roll_number)

# Function to calculate attendance for a given subject
def calculate_subject_attendance(roll_number, subject):
    # Find the row index of the specified roll number
    roll_numbers = [str(sheet.cell(row=i, column=1).value) for i in range(1, sheet.max_row + 1)]
    if str(roll_number) not in roll_numbers:
        print("Roll number not found!")
        return None
    else:
        roll_index = roll_numbers.index(str(roll_number)) + 1  # Adding 1 to match Excel row index

    # Find the column index of the specified subject
    subject_column_mapping = {
        'AOA': 3,
        'OS': 4,
        'DBMS': 5,
        'MP': 6,
        'M4': 7,
        'PYTHON': 8
    }
    if subject.upper() not in subject_column_mapping:
        print("Subject not found!")
        return None
    else:
        subject_column_index = subject_column_mapping[subject.upper()]

    # Get attendance for the specified subject
    attendance = sheet.cell(row=roll_index, column=subject_column_index).value
    return attendance

# Function to calculate average attendance for all subjects
def calculate_average_attendance(roll_number):
    # Find the row index of the specified roll number
    roll_numbers = [str(sheet.cell(row=i, column=1).value) for i in range(1, sheet.max_row + 1)]
    if str(roll_number) not in roll_numbers:
        print("Roll number not found!")
        return None
    else:
        roll_index = roll_numbers.index(str(roll_number)) + 1  # Adding 1 to match Excel row index

    # Get attendance data for all subjects
    attendance_data = [sheet.cell(row=roll_index, column=i).value for i in range(3, 9)]  # Assuming attendance data is in columns C to H
    total_attendance = sum(attendance for attendance in attendance_data if attendance is not None)
    num_subjects = len(attendance_data)
    if num_subjects == 0:
        return 0  # No attendance data found
    else:
        return total_attendance / num_subjects

# Function to send WhatsApp message
def send_whatsapp_message(phone_number, message):
    # Get the current time
    now = datetime.datetime.now()
    time_hour = now.hour
    time_min = now.minute + 1  # Change to the desired time offset

    # Send the message
    pywhatkit.sendwhatmsg(phone_number, message, time_hour, time_min)
    print(f"Message sent to {phone_number}")

# Prompt the user to input the roll number
roll_number = input("Enter roll number: ")

# Prompt the user to choose an option
print("Choose an option:")
print("1. Calculate attendance for a specific subject")
print("2. Calculate average attendance for all subjects")
option = input("Enter your choice (1 or 2): ")

if option == '1':
    subject = input("Enter the subject (AOA, OS, DBMS, MP, M4, PYTHON): ")
    attendance = calculate_subject_attendance(roll_number, subject)
    if attendance is not None:
        phone_number = get_phone_number(int(roll_number))
        if phone_number:
            message = f"Your attendance in {subject.upper()} is {attendance}%"
            send_whatsapp_message(phone_number, message)
        else:
            print("Phone number not found for the given roll number.")
elif option == '2':
    average_attendance = calculate_average_attendance(roll_number)
    if average_attendance is not None:
        phone_number = get_phone_number(int(roll_number))
        if phone_number:
            message = f"Your average attendance across all subjects is {average_attendance:.2f}%"
            send_whatsapp_message(phone_number, message)
        else:
            print("Phone number not found for the given roll number.")
else:
    print("Invalid option!")

# Close the workbook after use
workbook.close()
